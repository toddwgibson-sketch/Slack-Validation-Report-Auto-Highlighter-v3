"""
Slack Validation Report Auto-Highlighter v3
============================================
Output matches the mismatch example format exactly:
- L&R columns next to T0 and T1 device columns
- Per-column colour coding (Source=pink, DMARC1=yellow, DMARC2=green, Dest=blue, Z=light blue, Active=red, Expected=green)
- Possible patch panel + Possible DMARC columns (from cutsheet reverse lookup)
- Physical rows = yellow, logical rows = light purple
- Three tabs: All | Mispatches | Downlinks
- Optics and FEC tabs also included
"""

import sys, os, re, copy, json, time
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
    HAS_TK = True
except ImportError:
    HAS_TK = False

CONFIG_FILE = Path.home() / ".highlight_slack_config.json"

def load_config():
    if CONFIG_FILE.exists():
        try: return json.loads(CONFIG_FILE.read_text())
        except: pass
    return {}

def save_config(cfg):
    CONFIG_FILE.write_text(json.dumps(cfg, indent=2))

def pick_file(title, filetypes=None):
    if not HAS_TK:
        path = input(f"{title}\nEnter file path: ").strip().strip('"').strip("'")
        return path if os.path.isfile(path) else None
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    path = filedialog.askopenfilename(
        title=title,
        filetypes=filetypes or [("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    root.destroy()
    return path or None

def pick_multiple_files(title, filetypes=None):
    """Pick one or more files — returns list of paths"""
    if not HAS_TK:
        print(f"{title}")
        print("Enter file paths one per line, blank line when done:")
        paths = []
        while True:
            p = input("  Path: ").strip().strip('"').strip("'")
            if not p: break
            if os.path.isfile(p): paths.append(p)
            else: print(f"  Not found: {p}")
        return paths
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    paths = filedialog.askopenfilenames(
        title=title,
        filetypes=filetypes or [("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    root.destroy()
    return list(paths) if paths else []

def show_msg(title, msg, error=False):
    print(f"{'ERROR: ' if error else ''}{msg}")
    if HAS_TK:
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        (messagebox.showerror if error else messagebox.showinfo)(title, msg)
        root.destroy()

# ── Colours matching example file exactly ─────────────────────────────────────
WHITE      = "FFFFFF"    # white
YELLOW     = "FFFF00"    # physical row highlight
LOG_BG     = "FFFFFF"    # logical row - white
GREEN_DONE = "92D050"    # completed rows
SRC_BG     = "FCE4D6"    # source port - pink
D1_BG      = "FFF2CC"    # DMARC1 - yellow
D2_BG      = "E2F0D9"    # DMARC2 - green
DEST_BG    = "D9EAF7"    # destination - blue
Z_BG       = "DDEBF7"    # Z device - light blue
ACT_BG     = "FFC7CE"    # active - red
EXP_BG     = "C6EFCE"    # expected - green
LR_BG      = "FFFFFF"    # L&R col on physical rows - white
LR_LOG     = "FFFFFF"    # L&R col on logical rows - white
HDR_BG     = "1F4E79"    # header - navy
HDR_FG     = "FFFFFF"
PP_BG      = "FCE4D6"    # possible patch panel - pink
PD_BG      = "FFF2CC"    # possible DMARC - yellow

TAB_ALL    = "1F4E79"
TAB_MISS   = "C00000"
TAB_DOWN   = "ED7D31"
TAB_OPT    = "833C00"
TAB_FEC    = "7030A0"

def fill(h):   return PatternFill("solid", fgColor=h)
def no_fill(): return PatternFill(fill_type=None)
def font(color="000000", bold=False, sz=9):
    return Font(bold=bold, color=color, name="Arial", size=sz)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=False)
def vcenter(): return Alignment(vertical="center", wrap_text=False)

# ── Cutsheet lookup ───────────────────────────────────────────────────────────
def _load_single_cutsheet(path, t0, t1, t1_rev):
    """Load one cutsheet — auto-detects column layout by header names."""
    wb = load_workbook(path, read_only=True)
    sheet = next((wb[n] for n in wb.sheetnames if 'installation' in n.lower()), wb[wb.sheetnames[0]])

    # Detect format from header row
    hdr = {str(sheet.cell(1,c).value or '').strip(): c for c in range(1, sheet.max_column+1)}

    # New format (Installation Sheet with named columns)
    if 'DeviceA Name' in hdr:
        c_t0h  = hdr['DeviceA Name'] - 1       # T0 hostname
        c_t0i  = hdr['DeviceA Port'] - 1        # T0 iface
        c_rka  = hdr['RackA'] - 1               # T0 rack
        c_lbl  = hdr['DeviceA Physical Port'] - 1  # T0 L&R
        c_src  = hdr['Source_port'] - 1          # source PP
        c_dst  = hdr['Destination_port'] - 1     # dest PP
        c_t1lbl= hdr.get('DeviceB Physical Port', hdr.get('DeviceB Physical', 0)) - 1
        c_t1h  = hdr['DeviceB Name'] - 1         # T1 hostname
        c_t1i  = hdr['DeviceB Port'] - 1         # T1 iface
        c_rkb  = hdr['RackB'] - 1               # T1 rack
        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[c_t0h]: continue
            t0h  = str(row[c_t0h] or '').strip()
            t0i  = str(row[c_t0i] or '').strip()
            lbl  = str(row[c_lbl]  or '').strip()
            rka  = str(row[c_rka]  or '').strip()
            src  = str(row[c_src]  or '').strip()
            dst  = str(row[c_dst]  or '').strip()
            t1lbl= str(row[c_t1lbl] or '').strip() if c_t1lbl >= 0 else ''
            t1h  = str(row[c_t1h]  or '').strip()
            t1i  = str(row[c_t1i]  or '').strip()
            rkb  = str(row[c_rkb]  or '').strip()
            if t0h and t0i and lbl and re.match(r'\d+[LR]$', lbl):
                k = (t0h, t0i)
                t0[k] = lbl; t1[k] = t1lbl
            if t1h and t1i:
                t1_rev[(t1h, t1i)] = {
                    'device_a':    f"{t0h} {t0i}",
                    't0_lbl':      lbl,
                    'rack_a':      rka,
                    'source_port': src,
                    'dmarc1':      dst,
                    'dmarc2':      '',
                }
                count += 1
        wb.close()
        return count

    # Legacy format (Sheet1-style: DeviceA truncated in col1)
    count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 11: continue
        lbl    = str(row[0] or '').strip()
        dev_a  = str(row[1] or '').strip()
        rack_a = str(row[2] or '').strip()
        src    = str(row[3] or '').strip()
        dmarc1 = str(row[4] or '').strip()
        dmarc2 = str(row[5] or '').strip()
        dev_b  = str(row[7] or '').strip()
        t1l    = str(row[10] or '').strip()
        if dev_a and lbl and re.match(r'\d+[LR]$', lbl):
            parts = dev_a.split()
            if len(parts) == 2:
                k = (parts[0], parts[1])
                t0[k] = lbl; t1[k] = t1l
        if dev_b and ' ' in dev_b:
            parts = dev_b.split()
            if len(parts) == 2:
                t1_rev[(parts[0], parts[1])] = {
                    'device_a':     dev_a,
                    't0_lbl':       lbl,
                    'rack_a':       rack_a,
                    'source_port':  src,
                    'dmarc1':       dmarc1,
                    'dmarc2':       dmarc2,
                }
                count += 1
    wb.close()
    return count

# Global PP lookup from cutsheet — fallback for logical rows with no physical partner in report
_cutsheet_pp = {}
# Global T1-side physical port label lookup: (t1_host, t1_iface) -> T1 L&R label
_t1_label_map = {}

def build_lookup(paths):
    """Build merged lookup from one or more cutsheet paths"""
    global _cutsheet_pp
    _cutsheet_pp = {}
    if isinstance(paths, str):
        paths = [paths]
    t0, t1, t1_rev = {}, {}, {}
    for path in paths:
        count = _load_single_cutsheet(path, t0, t1, t1_rev)
        # Also build PP data lookup keyed by (hostname, iface)
        wb2 = load_workbook(path, read_only=True)
        sheet2 = next((wb2[n] for n in wb2.sheetnames if 'installation' in n.lower()), wb2[wb2.sheetnames[0]])
        hdr2 = {str(sheet2.cell(1,c).value or '').strip(): c for c in range(1, sheet2.max_column+1)}
        if 'DeviceA Name' in hdr2:
            # New format — use named columns
            c_t0h = hdr2['DeviceA Name'] - 1
            c_t0i = hdr2['DeviceA Port'] - 1
            c_src = hdr2['Source_port'] - 1
            c_dst = hdr2['Destination_port'] - 1
            c_t1h2 = hdr2.get('DeviceB Name', 0) - 1
            c_t1i2 = hdr2.get('DeviceB Port', 0) - 1
            c_t1lr = hdr2.get('DeviceB Physical Port', 0) - 1
            for row in sheet2.iter_rows(min_row=2, values_only=True):
                if not row or not row[c_t0h]: continue
                t0h = str(row[c_t0h] or '').strip()
                t0i = str(row[c_t0i] or '').strip()
                if t0h and t0i:
                    _cutsheet_pp[(t0h, t0i)] = {
                        'source_port': str(row[c_src] or '').strip(),
                        'dmarc1':      str(row[c_dst] or '').strip(),
                        'dmarc2':      '',
                        'dest_port':   '',
                    }
                # T1-side label lookup
                if c_t1h2 >= 0 and c_t1i2 >= 0 and c_t1lr >= 0:
                    t1h2 = str(row[c_t1h2] or '').strip()
                    t1i2 = str(row[c_t1i2] or '').strip()
                    t1lr = str(row[c_t1lr] or '').strip()
                    if t1h2 and t1i2 and t1lr:
                        _t1_label_map[(t1h2, t1i2)] = t1lr
        else:
            # Legacy format
            for row in sheet2.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 7: continue
                dev_a = str(row[1] or '').strip()
                if dev_a and ' ' in dev_a:
                    parts = dev_a.split()
                    if len(parts) == 2:
                        _cutsheet_pp[(parts[0], parts[1])] = {
                            'source_port': str(row[3] or '').strip(),
                            'dmarc1':      str(row[4] or '').strip(),
                            'dmarc2':      str(row[5] or '').strip(),
                            'dest_port':   str(row[6] or '').strip(),
                        }
        wb2.close()
        print(f"    Loaded: {os.path.basename(path)} ({count} T1 entries)")
    return t0, t1, t1_rev

def get_prev_issues(report_path):
    """Extract all issues from a previous report for recurring detection"""
    try:
        wb = load_workbook(report_path, read_only=True)
    except Exception as e:
        print(f"  Warning: could not load previous report: {e}")
        return set(), set(), set()

    # LLDP issues
    ws = next((wb[n] for n in wb.sheetnames if 'lldp' in n.lower()), None)
    prev_miss = set(); prev_down = set(); prev_rack_map = {}
    if ws:
        hc = next((c for c in range(1,ws.max_column+1) if str(ws.cell(1,c).value or '').strip()=='Hostname'), None)
        ic = next((c for c in range(1,ws.max_column+1) if str(ws.cell(1,c).value or '').strip()=='Interface'), None)
        ac = next((c for c in range(1,ws.max_column+1) if str(ws.cell(1,c).value or '').strip() in ('Act. Interface','Act.Interface')), None)
        rc = next((c for c in range(1,ws.max_column+1) if str(ws.cell(1,c).value or '').strip()=='Rack'), None)
        if hc and ic and ac:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row: continue
                h=str(row[hc-1] or '').strip(); i=str(row[ic-1] or '').strip()
                ai=str(row[ac-1] or '').strip().lower()
                rack=str(row[rc-1] or '').strip() if rc else 'Unknown'
                if h and i:
                    if ai == 'interface down': prev_down.add((h,i))
                    elif ai.startswith('swp'):  prev_miss.add((h,i))
                    prev_rack_map[(h,i)] = rack or 'Unknown' 

    # Optics issues
    ws_opt = next((wb[n] for n in wb.sheetnames if 'optic' in n.lower()), None)
    prev_opt = set(); prev_opt_rack_map = {}
    if ws_opt:
        hc = next((c for c in range(1,ws_opt.max_column+1) if str(ws_opt.cell(1,c).value or '').strip()=='Hostname'), None)
        ic = next((c for c in range(1,ws_opt.max_column+1) if str(ws_opt.cell(1,c).value or '').strip() in ('Interface','Transceiver')), None)
        rc = next((c for c in range(1,ws_opt.max_column+1) if str(ws_opt.cell(1,c).value or '').strip()=='Rack'), None)
        if hc and ic:
            for row in ws_opt.iter_rows(min_row=2, values_only=True):
                if not row: continue
                h=str(row[hc-1] or '').strip(); i=str(row[ic-1] or '').strip()
                rack=str(row[rc-1] or '').strip() if rc else 'Unknown'
                if h and i:
                    prev_opt.add((h,i))
                    prev_opt_rack_map[(h,i)] = rack or 'Unknown'

    wb.close()
    print(f"  Previous report: {len(prev_miss)} mismatches, {len(prev_down)} downlinks, {len(prev_opt)} optics")
    return prev_miss, prev_down, prev_opt, prev_rack_map, prev_opt_rack_map

def get_history_flag(host, iface, current_type, prev_miss, prev_down, prev_opt):
    """Return (flag_text, flag_colour) for a row based on previous report"""
    key = (host, iface)
    if current_type == 'mismatch':
        if key in prev_miss: return "🔁 Recurring mismatch",  "FF6B6B"
        if key in prev_down: return "⬆️ Was downlink",        "FFB347"
    elif current_type == 'downlink':
        if key in prev_down: return "🔁 Recurring downlink",  "FF6B6B"
        if key in prev_opt:  return "⚡ Was optic error",      "D35400"   # orange-red — likely bad reseat
        if key in prev_miss: return "⬇️ Was mismatch",        "FFB347"
    elif current_type == 'optic':
        if key in prev_opt:  return "🔁 Recurring optic",     "FF6B6B"
        if key in prev_down: return "⬆️ Was downlink",        "FFB347"
        if key in prev_miss: return "⬇️ Was mismatch",        "FFB347"
    return "", ""

def get_labels(hostname, iface, phys_t0, phys_t1):
    key = (hostname, iface)
    if key in phys_t0:
        return phys_t0[key], phys_t1[key], True
    m = re.match(r'(swp\d+)s(\d+)', str(iface))
    if m:
        base, lane = m.group(1), int(m.group(2))
        partner_lane = {0:1, 1:0, 2:3, 3:2}.get(lane)
        if partner_lane is not None:
            p = (hostname, f"{base}s{partner_lane}")
            if p in phys_t0:
                return phys_t0[p], phys_t1[p], False
    return '', '', False

def row_type(act_iface):
    v = str(act_iface or '').strip().lower()
    if v == 'interface down': return 'downlink'
    if v.startswith('swp'):   return 'mismatch'
    return 'other'

def find_col(ws, *names):
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value or '').strip() in names:
            return c
    return None

# ── Build output sheet ────────────────────────────────────────────────────────
def build_lldp_sheet(wb_out, sheet_name, rows, tab_colour, is_mismatch=False,
                     prev_miss=None, prev_down=None, prev_opt=None, is_downlinks=False):
    prev_miss = prev_miss or set()
    prev_down = prev_down or set()
    prev_opt  = prev_opt  or set()
    ws = wb_out.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = tab_colour

    # Output column layout (matching example exactly):
    # 1=Hostname, 2=Interface, 3=L&R(T0), 4=Rack, 5=Elevation
    # 6=Source_port, 7=DMARC1, 8=DMARC2, 9=Destination_port
    # 10=Z Hostname, 11=Z Interface, 12=L&R(T1), 13=Z Rack, 14=Z Elevation
    # 15=Possible DMARC, 16=Possible patch panel  (mismatch cols)
    # 17=Active Host, 18=Act. Interface, 19=Act. Rack, 20=Act. Elevation
    # 21=Expected Hostname, 22=Exp. Interface, 23=Exp. Rack, 24=Exp. Elevation

    # Detect if DMARC2/Destination_port columns are used (JBP15-style 4-PP path)
    # JBP19-style has only T0 PP (cells[4]) and T1 PP (cells[5]), DMARC2/Dest empty
    has_dmarc = any(r['cells'][6]['value'] or r['cells'][7]['value'] for r in rows)
    if has_dmarc:
        pp_headers = [
            ("Source_port",          "C0504D"),
            ("DMARC1",               "7F6000"),
            ("DMARC2",               "375623"),
            ("Destination_port",     "17375E"),
        ]
        possible_pp_headers = [
            ("Possible Source Port", "833C00"),
            ("Possible DMARC1",      "7F6000"),
            ("Possible DMARC2",      "C0504D"),
        ]
        possible_t1_headers = []
    else:
        pp_headers = [
            ("T0 PP",                "C0504D"),
            ("T1 PP",                "17375E"),
        ]
        possible_pp_headers = [
            ("Possible T0 PP",       "833C00"),
            ("Possible T1 PP",       "C0504D"),
        ]
        possible_t1_headers = [
            ("Act. T1 Host",         "9C0006"),
            ("Act. T1 Interface",    "9C0006"),
            ("Act. T1 L&R",          "9C0006"),
            ("Act. T1 Rack",         "9C0006"),
            ("Act. T1 Elevation",    "9C0006"),
        ]

    base_headers = [
        ("Hostname",             HDR_BG),
        ("Interface",            HDR_BG),
        ("L&R",                  HDR_BG),
        ("Rack",                 HDR_BG),
        ("Elevation",            HDR_BG),
    ] + pp_headers + [
        ("Z Hostname",           "17375E"),
        ("Z Interface",          "17375E"),
        ("L&R",                  "17375E"),
        ("Z Rack",               "17375E"),
        ("Z Elevation",          "17375E"),
    ]
    possible_headers = [
        ("Possible Device A",    "833C00"),
        ("Possible Rack / U",    "833C00"),
    ] + possible_pp_headers + possible_t1_headers
    if is_downlinks or not has_dmarc:
        tail_headers = []
    else:
        tail_headers = [
            ("Active Host",          "9C0006"),
            ("Act. Interface",       "9C0006"),
            ("Act. Rack",            "9C0006"),
            ("Act. Elevation",       "9C0006"),
            ("Expected Hostname",    "375623"),
            ("Exp. Interface",       "375623"),
            ("Exp. Rack",            "375623"),
            ("Exp. Elevation",       "375623"),
        ]
    tail_headers += [("History", "595959")]
    headers = base_headers + (possible_headers if is_mismatch else []) + tail_headers

    for col, (label, bg) in enumerate(headers, start=1):
        c = ws.cell(1, col)
        c.value     = label
        c.fill      = fill(bg)
        c.font      = font(HDR_FG, bold=True, sz=9)
        c.alignment = center()

    ws.row_dimensions[1].height = 20

    # Column widths - possible cols only on mismatch tab
    base_widths     = [26,12,6,7,6,30,28,28,30,26,12,6,7,6]
    possible_widths = [8,14,30,28,28]
    tail_widths     = ([] if (is_downlinks or not has_dmarc) else [26,12,7,6,26,12,7,6]) + [22]
    widths = base_widths + (possible_widths if is_mismatch else []) + tail_widths
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for out_row, rd in enumerate(rows, start=2):
        p  = rd['is_phys']
        mi = rd.get('mismatch_info', {})

        # Row background
        row_bg = YELLOW if p else LOG_BG

        # Values mapping from row data
        cells = rd['cells']  # original cols 1-20 from lldp_sp
        # Original: 1=Hostname, 2=Interface, 3=Rack, 4=Elevation, 5=Source_port
        #           6=DMARC1, 7=DMARC2, 8=Destination_port
        #           9=Z Hostname, 10=Z Interface, 11=Z Rack, 12=Z Elevation
        #           13=Active Host, 14=Act.Interface, 15=Act.Rack, 16=Act.Elevation
        #           17=Expected Hostname, 18=Exp.Interface, 19=Exp.Rack, 20=Exp.Elevation

        na = '#N/A' if not p else ''
        if has_dmarc:
            pp_vals_row = [cells[4]['value'], cells[5]['value'], cells[6]['value'], cells[7]['value']]
            poss_pp_vals = [
                mi.get('source_port', na) if p else na,
                mi.get('dmarc1',      na) if p else na,
                mi.get('dmarc2',      na) if p else na,
            ]
        else:
            pp_vals_row  = [cells[4]['value'], cells[5]['value']]
            poss_pp_vals = [
                mi.get('source_port', na) if p else na,
                mi.get('dmarc1',      na) if p else na,  # holds T1 PP in new format
            ]
        base_values = [
            cells[0]['value'],   # Hostname
            cells[1]['value'],   # Interface
            rd['t0'],            # L&R T0
            cells[2]['value'],   # Rack
            cells[3]['value'],   # Elevation
        ] + pp_vals_row + [
            cells[8]['value'],   # Z Hostname
            cells[9]['value'],   # Z Interface
            rd['t1'],            # L&R T1
            cells[10]['value'],  # Z Rack
            cells[11]['value'],  # Z Elevation
        ]
        if is_mismatch:
            act_t1_vals = ([
                cells[12]['value'] or na,       # Act. T1 Host
                cells[13]['value'] or na,       # Act. T1 Interface
                rd.get('act_t1_lr', '') or na,  # Act. T1 L&R
                cells[14]['value'] or na,       # Act. T1 Rack
                cells[15]['value'] or na,       # Act. T1 Elevation
            ] if not has_dmarc else [])
            possible_values = [
                mi.get('t0_lbl', na) if p else na,
                mi.get('rack_a', na) if p else na,
            ] + poss_pp_vals + act_t1_vals
        else:
            possible_values = []
        # Get history flag
        host_val  = cells[0]['value']
        iface_val = cells[1]['value']
        hist_flag, hist_col = get_history_flag(
            str(host_val or '').strip(), str(iface_val or '').strip(),
            rd['row_type'], prev_miss, prev_down, prev_opt
        )
        if is_downlinks or not has_dmarc:
            tail_values = [hist_flag]
        else:
            tail_values = [
                cells[12]['value'],  # Active Host
                cells[13]['value'],  # Act. Interface
                cells[14]['value'],  # Act. Rack
                cells[15]['value'],  # Act. Elevation
                cells[16]['value'],  # Expected Hostname
                cells[17]['value'],  # Exp. Interface
                cells[18]['value'],  # Exp. Rack
                cells[19]['value'],  # Exp. Elevation
                hist_flag,
            ]
        all_values = base_values + possible_values + tail_values
        values = {i+1: v for i, v in enumerate(all_values)}

        # Physical rows: yellow across all base+tail cols, possible cols keep own colour
        # Logical rows: LOG_BG across all base+tail cols
        if p:
            _pp_fills = ["FFFFFF"]*4 if has_dmarc else ["FFFFFF"]*2
            base_fills = ["FFFFFF","FFFFFF",LR_BG,"FFFFFF","FFFFFF"] + _pp_fills + ["FFFFFF","FFFFFF",LR_BG,"FFFFFF","FFFFFF"]
            tail_fills = ["FFFFFF"]*8
        else:
            _pp_fills_l = [LOG_BG]*4 if has_dmarc else [LOG_BG]*2
            base_fills = [LOG_BG,LOG_BG,LR_LOG,LOG_BG,LOG_BG] + _pp_fills_l + [LOG_BG,LOG_BG,LR_LOG,LOG_BG,LOG_BG]
            tail_fills = [LOG_BG]*8
        if is_mismatch:
            if has_dmarc:
                possible_fills = ["FDDCB5","FDDCB5","FDDCB5", D1_BG, PP_BG]
            else:
                # Device A, Rack, T0 PP, T1 PP, Act. T1 Host, Interface, Rack, Elevation
                possible_fills = ["FDDCB5","FDDCB5","FDDCB5","FDDCB5","FDDCB5","FDDCB5","FDDCB5","FDDCB5","FDDCB5"]
        else:
            possible_fills = []
        col_fills = {i+1: f for i, f in enumerate(base_fills + possible_fills + tail_fills)}
        for _c in range(1, len(all_values)+1):
            if _c not in col_fills: col_fills[_c] = row_bg

        for col in range(1, len(all_values) + 1):
            c = ws.cell(out_row, col)
            c.value     = values.get(col, '')
            c.fill      = fill(col_fills[col])
            c.font      = font(sz=8)
            c.alignment = vcenter()

        ws.row_dimensions[out_row].height = 15

    ws.freeze_panes = "A2"

    # ── Draw borders around each physical+logical pair ────────────────────────
    from openpyxl.styles import Border, Side
    thin  = Side(style="thin",   color="AAAAAA")
    thick = Side(style="medium", color="555555")

    total_data_cols = len(all_values) if rows else 0
    if total_data_cols == 0 and ws.max_column > 1:
        total_data_cols = ws.max_column

    # Group consecutive rows by their L&R value (col 3 = L&R after T0 port col)
    lr_col = 3  # L&R is col 3 in output (col1=Hostname, col2=Interface, col3=L&R)
    data_row = 2
    max_r = ws.max_row

    while data_row <= max_r:
        lr_val = ws.cell(data_row, lr_col).value
        # Find how many consecutive rows share the same L&R value
        group_end = data_row
        while group_end + 1 <= max_r and ws.cell(group_end+1, lr_col).value == lr_val and lr_val:
            group_end += 1

        # Apply border box around this group
        for row in range(data_row, group_end + 1):
            is_top    = (row == data_row)
            is_bottom = (row == group_end)
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                cell.border = Border(
                    top    = thick if is_top    else thin,
                    bottom = thick if is_bottom else Side(style=None),
                    left   = thick if col == 1  else thin,
                    right  = thick if col == ws.max_column else thin,
                )

        data_row = group_end + 1

# ── Process source sheet ──────────────────────────────────────────────────────
def read_lldp_rows(ws_src, phys_t0, phys_t1, t1_rev):
    host_col    = find_col(ws_src, 'Hostname')
    iface_col   = find_col(ws_src, 'Interface')
    rack_col    = find_col(ws_src, 'Rack')
    elev_col    = find_col(ws_src, 'Elevation')
    act_h_col   = find_col(ws_src, 'Active Host')
    act_if_col  = find_col(ws_src, 'Act. Interface', 'Act.Interface')
    act_rack_col= find_col(ws_src, 'Act. Rack', 'Act.Rack')
    act_elev_col= find_col(ws_src, 'Act. Elevation', 'Act.Elevation')
    exp_h_col   = find_col(ws_src, 'Expected Hostname', 'Exp. Hostname')
    exp_if_col  = find_col(ws_src, 'Exp. Interface', 'Exp.Interface')
    exp_rack_col= find_col(ws_src, 'Exp. Rack', 'Exp.Rack')
    exp_elev_col= find_col(ws_src, 'Exp. Elevation', 'Exp.Elevation')
    z_h_col     = find_col(ws_src, 'Z Hostname', 'Z Host')
    z_if_col    = find_col(ws_src, 'Z Interface')
    z_rack_col  = find_col(ws_src, 'Z Rack')
    z_elev_col  = find_col(ws_src, 'Z Elevation')
    src_col     = find_col(ws_src, 'Source_port', 'Source port')
    d1_col      = find_col(ws_src, 'DMARC1')
    d2_col      = find_col(ws_src, 'DMARC2')
    dst_col     = find_col(ws_src, 'Destination_port', 'Destination port')
    ncols = ws_src.max_column


    # Single pass — cache all row data first (read_only worksheets only allow one iteration)
    all_rows_data = {}  # row_num -> {col: value}
    col_indices = [c for c in [host_col, iface_col, rack_col, elev_col,
                                src_col, d1_col, d2_col, dst_col,
                                z_h_col, z_if_col, z_rack_col, z_elev_col,
                                act_h_col, act_if_col, act_rack_col, act_elev_col,
                                exp_h_col, exp_if_col, exp_rack_col, exp_elev_col]
                   if c and c > 0]
    for row_data in ws_src.iter_rows(min_row=2, values_only=False):
        if not row_data: continue
        rn = row_data[0].row
        all_rows_data[rn] = {c: row_data[c-1].value for c in col_indices if c <= len(row_data)}

    def rget(rn, col):
        if not col or col <= 0: return None
        return all_rows_data.get(rn, {}).get(col)

    raw_rows = {}
    for rn, data in all_rows_data.items():
        h = str(data.get(host_col, '') or '').strip()
        i = str(data.get(iface_col, '') or '').strip()
        if h and i: raw_rows[(h, i)] = rn

    rows = []
    for row in sorted(all_rows_data.keys()):
        host  = str(rget(row, host_col)  or '').strip()
        iface = str(rget(row, iface_col) or '').strip()
        if not host or not iface: continue

        t0, t1, is_p = get_labels(host, iface, phys_t0, phys_t1)
        act_if = rget(row, act_if_col)
        rtype  = row_type(act_if)

        mi = {}
        if is_p and act_h_col and act_if_col:
            ah = str(rget(row, act_h_col) or '').strip()
            ai = str(rget(row, act_if_col) or '').strip()
            if ai.lower().startswith('swp'):
                mi = t1_rev.get((ah, ai), {})
                if not mi:
                    m2 = re.match(r'(swp\d+)s(\d+)', ai)
                    if m2:
                        base2, lane2 = m2.group(1), int(m2.group(2))
                        partner = {0:1,1:0,2:3,3:2}.get(lane2)
                        if partner is not None:
                            mi = t1_rev.get((ah, f"{base2}s{partner}"), {})
        # Build normalised cells list — consistent column order regardless of source format
        # Position: 0=Hostname, 1=Interface, 2=Rack, 3=Elevation,
        #           4=Source_port, 5=DMARC1, 6=DMARC2, 7=Destination_port,
        #           8=Z Hostname, 9=Z Interface, 10=Z Rack, 11=Z Elevation,
        #           12=Active Host, 13=Act. Interface, 14=Act. Rack, 15=Act. Elevation,
        #           16=Exp. Hostname, 17=Exp. Interface, 18=Exp. Rack, 19=Exp. Elevation
        def mkc(col): return {'value': rget(row, col), 'fill': no_fill()}
        # Z cols: use explicit Z cols if present, otherwise fall back to Expected cols
        # (JBP19 has no Z cols — Expected Hostname/Interface/Rack carry the T1 info)
        z_h   = z_h_col   or exp_h_col
        z_if  = z_if_col  or exp_if_col
        z_rk  = z_rack_col or exp_rack_col
        z_el  = z_elev_col or exp_elev_col
        cells = [
            mkc(host_col), mkc(iface_col), mkc(rack_col), mkc(elev_col),
            mkc(src_col),  mkc(d1_col),   mkc(d2_col),   mkc(dst_col),
            mkc(z_h),      mkc(z_if),     mkc(z_rk),     mkc(z_el),
            mkc(act_h_col), mkc(act_if_col), mkc(act_rack_col), mkc(act_elev_col),
            mkc(exp_h_col), mkc(exp_if_col), mkc(exp_rack_col), mkc(exp_elev_col),
        ]

        # cells always has exactly 20 normalised entries

        # For physical rows — fill PP from cutsheet if not in report
        if is_p and not cells[4]['value']:
            pp_data = _cutsheet_pp.get((host, iface), {})
            if pp_data:
                for pp_col, key in [(4,'source_port'),(5,'dmarc1'),(6,'dmarc2'),(7,'dest_port')]:
                    cells[pp_col] = {'value': pp_data.get(key,''), 'fill': no_fill()}

        # For logical rows — copy patch panel cols from physical partner
        # First try partner in the report, then fall back to cutsheet
        # Patch panel cols (0-indexed): 4=Source_port, 5=DMARC1, 6=DMARC2, 7=Destination_port
        if not is_p:
            m = re.match(r'(swp\d+)s(\d+)', iface)
            if m:
                base, lane = m.group(1), int(m.group(2))
                partner_lane = {0:1,1:0,2:3,3:2}.get(lane)
                partner_iface = f"{base}s{partner_lane}"
                partner_row = raw_rows.get((host, partner_iface))
                if partner_row:
                    # Try cutsheet PP first, then raw partner row cells
                    pp_data = _cutsheet_pp.get((host, partner_iface), {})
                    if pp_data:
                        for pp_col, key in [(4,'source_port'),(5,'dmarc1'),(6,'dmarc2'),(7,'dest_port')]:
                            cells[pp_col] = {'value': pp_data.get(key,''), 'fill': no_fill()}
                    elif src_col:
                        for idx, col in enumerate([src_col, d1_col, d2_col, dst_col]):
                            if col:
                                cells[4+idx] = {'value': rget(partner_row, col), 'fill': no_fill()}
                else:
                    # Partner not in report — fall back to cutsheet PP data
                    pp_data = _cutsheet_pp.get((host, partner_iface), {})
                    if pp_data:
                        for pp_col, key in [(4,'source_port'),(5,'dmarc1'),(6,'dmarc2'),(7,'dest_port')]:
                            cells[pp_col] = {'value': pp_data.get(key,''), 'fill': no_fill()}

        # Enrich t1 label from T1-side label map if not already set
        z_host_val  = cells[8]['value']
        z_iface_val = cells[9]['value']
        t1_enriched = t1 or _t1_label_map.get((str(z_host_val or ''), str(z_iface_val or '')), '')

        # Act. T1 L&R label
        act_t1_h  = str(cells[12]['value'] or '')
        act_t1_if = str(cells[13]['value'] or '')
        act_t1_lr = _t1_label_map.get((act_t1_h, act_t1_if), '')

        rows.append({
            't0': t0, 't1': t1_enriched, 'is_phys': is_p,
            'row_type': rtype, 'cells': cells,
            'mismatch_info': mi,
            'act_t1_lr': act_t1_lr,
        })
    return rows


# ── Summary Tab ───────────────────────────────────────────────────────────────
def build_summary_tab(wb_out, lldp_rows, miss_rows, down_rows,
                      prev_miss, prev_down, prev_opt,
                      report_name, prev_report_name,
                      prev_rack_map=None, prev_opt_rack_map=None, curr_opt_rack=None):
    prev_rack_map     = prev_rack_map     or {}
    prev_opt_rack_map = prev_opt_rack_map or {}
    curr_opt_rack     = curr_opt_rack     or {}  # rack -> set of (host,iface) current optics
    import re as _re
    from datetime import datetime

    ws = wb_out.create_sheet("Summary", 0)  # insert as first tab
    ws.sheet_properties.tabColor = "1F4E79"

    NAVY  = "1F4E79"; WHITE = "FFFFFF"; RED   = "C00000"
    GREEN = "1E8449"; AMBER = "B7770D"; TEAL  = "0D7377"
    LRED  = "FADBD8"; LGRN  = "D5F5E3"; LYEL  = "FEF9E7"
    LGRY  = "F2F2F2"; DGRY  = "595959"; ORNG  = "E67E22"

    def fill(h):  return PatternFill("solid", fgColor=h)
    def font(color="000000", bold=False, sz=10, italic=False):
        return Font(bold=bold, italic=italic, color=color, name="Arial", size=sz)
    def center(wrap=False): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    def left():             return Alignment(horizontal="left",   vertical="center", wrap_text=False)

    # ── Classify each row ──────────────────────────────────────────────────────
    def classify(rows, current_type, pm, pd, po):
        total = new = recurring = type_change = 0
        for rd in rows:
            h = str(rd['cells'][0]['value'] or '').strip()
            i = str(rd['cells'][1]['value'] or '').strip()
            key = (h, i)
            flag, _ = get_history_flag(h, i, current_type, pm, pd, po)
            total += 1
            if not flag:                    new += 1
            elif '🔁' in flag:             recurring += 1
            else:                           type_change += 1
        return total, new, recurring, type_change

    miss_total,  miss_new,  miss_rec,  miss_tc  = classify(miss_rows,  'mismatch', prev_miss, prev_down, prev_opt)
    down_total,  down_new,  down_rec,  down_tc  = classify(down_rows,  'downlink', prev_miss, prev_down, prev_opt)

    # Optics from lldp_rows context not available directly — derive from tab
    # Use lldp_rows all for total counts
    total_issues = len(lldp_rows)
    has_prev = bool(prev_miss or prev_down or prev_opt)

    # Per-rack breakdown
    # Build current rack data with actual key sets for accurate Fixed calculation
    rack_data = {}
    curr_miss_by_rack = {}  # rack -> set of (host,iface) currently mismatch
    curr_down_by_rack = {}  # rack -> set of (host,iface) currently downlink

    curr_opt_by_rack = curr_opt_rack  # populated from optics sheet in main

    for rd in lldp_rows:
        h    = str(rd['cells'][0]['value'] or '').strip()
        rack = str(rd['cells'][2]['value'] or '').strip() or 'Unknown'
        if rack not in rack_data:
            rack_data[rack] = {'miss':0,'down':0,'opt':0,'miss_rec':0,'down_rec':0,'miss_new':0,'down_new':0}
            curr_miss_by_rack[rack] = set()
            curr_down_by_rack[rack] = set()
            if rack not in curr_opt_by_rack: curr_opt_by_rack[rack] = set()
        key = (str(rd['cells'][0]['value'] or '').strip(),
               str(rd['cells'][1]['value'] or '').strip())
        rtype = rd['row_type']
        flag, _ = get_history_flag(key[0], key[1], rtype, prev_miss, prev_down, prev_opt)
        if rtype == 'mismatch':
            rack_data[rack]['miss'] += 1
            curr_miss_by_rack[rack].add(key)
            if '🔁' in flag: rack_data[rack]['miss_rec'] += 1
            elif not flag:   rack_data[rack]['miss_new'] += 1
        elif rtype == 'downlink':
            rack_data[rack]['down'] += 1
            curr_down_by_rack[rack].add(key)
            if '🔁' in flag: rack_data[rack]['down_rec'] += 1
            elif not flag:   rack_data[rack]['down_new'] += 1

    # Build previous rack data key sets directly from prev_rack_map
    # prev_rack_map is passed in so we can look up rack for prev report keys
    # even if those links no longer exist in the current report
    prev_miss_by_rack = {}
    prev_down_by_rack = {}
    prev_opt_by_rack  = {}

    for (h, i), rack in prev_rack_map.items():
        key = (h, i)
        if key in prev_miss:
            if rack not in prev_miss_by_rack: prev_miss_by_rack[rack] = set()
            prev_miss_by_rack[rack].add(key)
        if key in prev_down:
            if rack not in prev_down_by_rack: prev_down_by_rack[rack] = set()
            prev_down_by_rack[rack].add(key)

    for (h, i), rack in prev_opt_rack_map.items():
        key = (h, i)
        if key in prev_opt:
            if rack not in prev_opt_by_rack: prev_opt_by_rack[rack] = set()
            prev_opt_by_rack[rack].add(key)

    # Also build current optics by rack from optics tab data
    # We need to pass current optics rows into summary - use lldp_rows for rack lookup
    # and cross-ref against all current optic keys
    # For now build from lldp_rows host_to_rack for any optic key we know about
    host_to_rack_curr = {(str(rd['cells'][0]['value'] or '').strip(),
                          str(rd['cells'][1]['value'] or '').strip()):
                         str(rd['cells'][2]['value'] or '').strip() or 'Unknown'
                         for rd in lldp_rows}

    # Store key sets in rack_data for fixed calculation
    all_racks_set = (set(rack_data.keys()) | set(prev_miss_by_rack.keys()) |
                     set(prev_down_by_rack.keys()) | set(prev_opt_by_rack.keys()) |
                     set(curr_opt_by_rack.keys()))
    for rack in all_racks_set:
        if rack not in rack_data:
            rack_data[rack] = {'miss':0,'down':0,'opt':0,'miss_rec':0,'down_rec':0,'miss_new':0,'down_new':0}
        rack_data[rack]['curr_miss_keys'] = curr_miss_by_rack.get(rack, set())
        rack_data[rack]['curr_down_keys'] = curr_down_by_rack.get(rack, set())
        rack_data[rack]['curr_opt_keys']  = curr_opt_by_rack.get(rack, set())  # from curr_opt_rack passed in
        rack_data[rack]['prev_miss_keys'] = prev_miss_by_rack.get(rack, set())
        rack_data[rack]['prev_down_keys'] = prev_down_by_rack.get(rack, set())
        rack_data[rack]['prev_opt_keys']  = prev_opt_by_rack.get(rack, set())

    # ── Layout ─────────────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 3   # left margin
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 3   # gap

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.merge_cells("B1:G1")
    c = ws["B1"]; c.value = "VALIDATION REPORT — SUMMARY"
    c.fill = fill(NAVY); c.font = Font(bold=True, color=WHITE, name="Arial", size=14)
    c.alignment = center(); ws.row_dimensions[1].height = 32

    ws.merge_cells("B2:G2")
    c = ws["B2"]; c.value = f"Report: {report_name}   |   Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.fill = fill(TEAL); c.font = Font(italic=True, color=WHITE, name="Arial", size=9)
    c.alignment = center(); ws.row_dimensions[2].height = 16

    if prev_report_name:
        ws.merge_cells("B3:G3")
        c = ws["B3"]; c.value = f"Compared against: {prev_report_name}"
        c.fill = fill("2E4057"); c.font = Font(italic=True, color=WHITE, name="Arial", size=9)
        c.alignment = center(); ws.row_dimensions[3].height = 14
    else:
        ws.merge_cells("B3:G3")
        c = ws["B3"]; c.value = "No previous report selected — recurring analysis not available"
        c.fill = fill(DGRY); c.font = Font(italic=True, color=WHITE, name="Arial", size=9)
        c.alignment = center(); ws.row_dimensions[3].height = 14

    # ── KPI Banner ─────────────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 8  # spacer

    kpi_labels = ["TOTAL ISSUES", "MISMATCHES", "DOWNLINKS",
                  "🔁 RECURRING", "🆕 NEW TODAY"]
    kpi_values = [
        total_issues,
        miss_total,
        down_total,
        miss_rec + down_rec,
        miss_new + down_new,
    ]
    kpi_bgs = [NAVY, RED, AMBER, "C00000", GREEN]

    for i, (lbl, val, bg) in enumerate(zip(kpi_labels, kpi_values, kpi_bgs)):
        col = i + 2  # B=2 through G=7
        ltr = chr(64 + col)
        ws.row_dimensions[5].height = 16
        ws.row_dimensions[6].height = 30
        c = ws.cell(5, col); c.value = lbl
        c.fill = fill(bg); c.font = Font(bold=True, color=WHITE, name="Arial", size=8)
        c.alignment = center(wrap=True)
        c = ws.cell(6, col); c.value = val
        c.fill = fill(bg); c.font = Font(bold=True, color=WHITE, name="Arial", size=20)
        c.alignment = center()

    # ── Error Type Breakdown ───────────────────────────────────────────────────
    ws.row_dimensions[7].height = 10  # spacer
    ws.merge_cells("B8:G8")
    c = ws["B8"]; c.value = "ERROR TYPE BREAKDOWN"
    c.fill = fill(NAVY); c.font = Font(bold=True, color=WHITE, name="Arial", size=10)
    c.alignment = center(); ws.row_dimensions[8].height = 20

    # Headers
    hdrs = ["Type", "Total", "🆕 New", "🔁 Recurring", "↔️ Type Change", "% Recurring"]
    bgs  = [NAVY, NAVY, GREEN, RED, ORNG, NAVY]
    for i, (h, bg) in enumerate(zip(hdrs, bgs)):
        c = ws.cell(9, i+2); c.value = h
        c.fill = fill(bg); c.font = Font(bold=True, color=WHITE, name="Arial", size=9)
        c.alignment = center(); ws.row_dimensions[9].height = 18

    rows_data = [
        ("Mispatches",  miss_total,  miss_new,  miss_rec,  miss_tc),
        ("Downlinks",   down_total,  down_new,  down_rec,  down_tc),
    ]
    for row_i, (lbl, tot, new_, rec, tc) in enumerate(rows_data):
        row = 10 + row_i
        ws.row_dimensions[row].height = 20
        pct = f"{round(rec/tot*100)}%" if tot > 0 else "—"
        row_bg = LRED if rec > 0 else LGRN
        vals = [lbl, tot, new_, rec, tc, pct]
        bgs2  = [LGRY, LGRY, LGRN, LRED, LYEL, LGRY]
        for col_i, (v, bg) in enumerate(zip(vals, bgs2)):
            c = ws.cell(row, col_i+2)
            c.value = v; c.fill = fill(bg)
            bold = col_i in (0, 1)
            c.font = Font(bold=bold, name="Arial", size=10,
                          color=RED if bg==LRED and v else (GREEN if bg==LGRN and v else "000000"))
            c.alignment = center() if col_i > 0 else left()

    # ── Per-Rack Breakdown ─────────────────────────────────────────────────────
    # Build previous report rack data
    prev_rack = {}
    for key in prev_miss:
        # key = (hostname, iface) — need rack from current lldp_rows
        for rd in lldp_rows:
            h2 = str(rd['cells'][0]['value'] or '').strip()
            i2 = str(rd['cells'][1]['value'] or '').strip()
            if (h2, i2) == key:
                r2 = str(rd['cells'][2]['value'] or '').strip() or 'Unknown'
                if r2 not in prev_rack: prev_rack[r2] = {'miss':0,'down':0,'opt':0}
                prev_rack[r2]['miss'] += 1
                break
    for key in prev_down:
        for rd in lldp_rows:
            h2 = str(rd['cells'][0]['value'] or '').strip()
            i2 = str(rd['cells'][1]['value'] or '').strip()
            if (h2, i2) == key:
                r2 = str(rd['cells'][2]['value'] or '').strip() or 'Unknown'
                if r2 not in prev_rack: prev_rack[r2] = {'miss':0,'down':0,'opt':0}
                prev_rack[r2]['down'] += 1
                break

    # All racks from either prev or current
    all_racks = sorted(set(list(rack_data.keys()) + list(prev_rack.keys())))

    ws.row_dimensions[13].height = 10  # spacer
    ws.merge_cells("B14:N14")
    c = ws["B14"]; c.value = "PER-RACK BREAKDOWN  —  Previous vs Now"
    c.fill = fill(NAVY); c.font = Font(bold=True, color=WHITE, name="Arial", size=10)
    c.alignment = center(); ws.row_dimensions[14].height = 20

    # Column layout:
    # B=Rack | C=Miss Prev | D=Miss Now | E=Miss Fixed | F=Miss New
    #        | G=Down Prev | H=Down Now | I=Down Fixed | J=Down New
    #        | K=Opt Prev  | L=Opt Now  | M=Opt Fixed  | N=Opt New
    ws.column_dimensions['B'].width = 10
    for ltr, w in zip('CDEFGHIJKLMN', [9,9,9,9, 9,9,9,9, 9,9,9,9]):
        ws.column_dimensions[ltr].width = w

    # Group headers row 15
    for col, label, bg in [
        (3,  "MISMATCHES", "C00000"),
        (7,  "DOWNLINKS",  AMBER),
        (11, "OPTICS",     "7D3C98"),
    ]:
        ws.merge_cells(start_row=15, start_column=col, end_row=15, end_column=col+3)
        c = ws.cell(15, col); c.value = label; c.fill = fill(bg)
        c.font = Font(bold=True, color=WHITE, name="Arial", size=9)
        c.alignment = center()
    ws.cell(15, 2).value = "Rack"; ws.cell(15, 2).fill = fill(NAVY)
    ws.cell(15, 2).font = Font(bold=True, color=WHITE, name="Arial", size=9)
    ws.cell(15, 2).alignment = center()
    ws.row_dimensions[15].height = 16

    # Sub-headers row 16
    sub_hdrs = ["Rack", "Prev","Now","Fixed","New", "Prev","Now","Fixed","New", "Prev","Now","Fixed","New"]
    sub_bgs  = [NAVY, LGRY,LGRY,LGRN,LRED, LGRY,LGRY,LGRN,LRED, LGRY,LGRY,LGRN,LRED]
    for i, (h2, bg) in enumerate(zip(sub_hdrs, sub_bgs)):
        c = ws.cell(16, i+2); c.value = h2; c.fill = fill(bg)
        c.font = Font(bold=True, color="000000" if i>0 else WHITE, name="Arial", size=8)
        c.alignment = center()
    ws.row_dimensions[16].height = 14

    # Data rows
    for row_i, rack in enumerate(all_racks):
        row = 17 + row_i
        ws.row_dimensions[row].height = 18
        curr = rack_data.get(rack, {'miss':0,'down':0,'miss_new':0,'down_new':0})
        prev = prev_rack.get(rack, {'miss':0,'down':0,'opt':0})

        curr_miss = curr.get('miss', 0)
        curr_down = curr.get('down', 0)
        curr_new_miss = curr.get('miss_new', 0)
        curr_new_down = curr.get('down_new', 0)
        prev_miss_r = len(curr.get('prev_miss_keys', set()))
        prev_down_r = len(curr.get('prev_down_keys', set()))

        # Fixed = links in prev report for this rack that are NOT in current report
        # Uses actual set of keys stored per rack rather than arithmetic
        miss_fixed = len(curr.get('prev_miss_keys', set()) - curr.get('curr_miss_keys', set()))
        down_fixed = len(curr.get('prev_down_keys', set()) - curr.get('curr_down_keys', set()))
        opt_fixed  = len(curr.get('prev_opt_keys',  set()) - curr.get('curr_opt_keys',  set()))

        curr_opt   = len(curr.get('curr_opt_keys', set()))
        prev_opt_r = len(curr.get('prev_opt_keys', set()))
        curr_new_opt = len(curr.get('curr_opt_keys', set()) - curr.get('prev_opt_keys', set()))

        vals = [
            rack,
            prev_miss_r, curr_miss,  miss_fixed, curr_new_miss,
            prev_down_r, curr_down,  down_fixed, curr_new_down,
            prev_opt_r,  curr_opt,   opt_fixed,  curr_new_opt,
        ]
        for col_i, v in enumerate(vals):
            c = ws.cell(row, col_i+2)
            c.value = v
            # Colour: Fixed=green, New=red, others plain
            if col_i in (3, 7, 11) and v > 0:   c.fill = fill(LGRN)  # fixed
            elif col_i in (4, 8, 12) and v > 0:  c.fill = fill(LRED)  # new
            else:                                  c.fill = fill("FFFFFF" if col_i>0 else LGRY)
            c.font = Font(bold=(col_i==0), name="Arial", size=9,
                          color=GREEN if col_i in (3,7,11) and v>0
                          else RED if col_i in (4,8,12) and v>0 else "000000")
            c.alignment = center() if col_i > 0 else left()

    print(f"  Summary tab built — {len(all_racks)} racks")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    cfg = load_config()

    # Step 1: cutsheets — support multiple for full mismatch coverage
    saved_paths = cfg.get('cutsheet_paths', [])
    # Back-compat: handle old single path config
    if not saved_paths and cfg.get('cutsheet_path'):
        saved_paths = [cfg['cutsheet_path']]
    saved_paths = [p for p in saved_paths if os.path.isfile(p)]

    cutsheet_paths = []
    if saved_paths:
        names = '\n'.join(f"  • {os.path.basename(p)}" for p in saved_paths)
        if HAS_TK:
            root = tk.Tk()
            root.withdraw()
            root.attributes('-topmost', True)
            use_saved = messagebox.askyesno("Physical Cutsheets",
                f"Use saved cutsheets?\n\n{names}\n\nClick No to pick different ones.")
            root.destroy()
        else:
            print(f"Saved cutsheets:\n{names}")
            use_saved = input("Use these? (y/n): ").strip().lower() != 'n'
        if use_saved:
            cutsheet_paths = saved_paths

    if not cutsheet_paths:
        show_msg("Select Cutsheets",
            "Select all physical cutsheets for this job.\n\nHold Ctrl to select multiple files.")
        cutsheet_paths = pick_multiple_files(
            "Select Physical Cutsheet(s) — hold Ctrl for multiple")
        if not cutsheet_paths:
            show_msg("Cancelled", "No cutsheets selected.", error=True); sys.exit(0)
        cfg['cutsheet_paths'] = cutsheet_paths
        cfg['cutsheet_path']  = cutsheet_paths[0]  # back-compat
        save_config(cfg)

    # Step 2: previous report (optional)
    time.sleep(0.3)
    prev_report_path = None
    if HAS_TK:
        root = tk.Tk(); root.withdraw(); root.attributes('-topmost', True)
        want_prev = messagebox.askyesno("Previous Report",
            "Do you have a previous report to compare against?\n\n"
            "This will flag recurring issues across all tabs.")
        root.destroy()
    else:
        want_prev = input("Compare against a previous report? (y/n): ").strip().lower() == 'y'

    if want_prev:
        time.sleep(0.3)
        prev_report_path = pick_file("Select PREVIOUS Slack Validation Report")
        if prev_report_path:
            print(f"Previous report: {os.path.basename(prev_report_path)}")

    # Step 3: current report
    time.sleep(0.4)
    report_path = pick_file("Select Slack Validation Report")
    if not report_path:
        show_msg("Cancelled", "No report selected.", error=True); sys.exit(0)

    # Check for local hall cutsheet in same folder - append to list
    local = os.path.join(os.path.dirname(report_path), "cutsheet.xlsx")
    if os.path.isfile(local) and local not in cutsheet_paths:
        cutsheet_paths.append(local)
        print(f"Also loading local hall cutsheet: {local}")

    print(f"Loading cutsheet...")
    phys_t0, phys_t1, t1_rev = build_lookup(cutsheet_paths)
    print(f"  {len(phys_t0)} T0 entries | {len(t1_rev)} T1 reverse entries")

    # Load previous report issues if provided
    prev_miss = set(); prev_down = set(); prev_opt = set(); prev_rack_map = {}; prev_opt_rack_map = {}
    if prev_report_path:
        print("Loading previous report for comparison...")
        prev_miss, prev_down, prev_opt, prev_rack_map, prev_opt_rack_map = get_prev_issues(prev_report_path)

    print(f"Processing: {os.path.basename(report_path)}")
    wb_src = load_workbook(report_path)

    def find_sheet(wb, *patterns):
        best = None; best_score = 0
        for name in wb.sheetnames:
            for p in patterns:
                if p.lower() in name.lower():
                    score = len(p)
                    if score > best_score:
                        best = wb[name]; best_score = score
        return best

    ws_lldp   = find_sheet(wb_src, 'lldp')
    ws_optics = find_sheet(wb_src, 'optics_rx_tx', 'rx_tx', 'optic')
    ws_fec    = find_sheet(wb_src, 'combined_fec', 'fec')

    wb_out = Workbook(); wb_out.remove(wb_out.active)

    lldp_rows = miss_rows = down_rows = []

    if ws_lldp:
        lldp_rows = read_lldp_rows(ws_lldp, phys_t0, phys_t1, t1_rev)
        miss_rows = [r for r in lldp_rows if r['row_type'] == 'mismatch']
        down_rows = [r for r in lldp_rows if r['row_type'] == 'downlink']

        build_lldp_sheet(wb_out, "Mispatches", miss_rows, TAB_MISS, is_mismatch=True, prev_miss=prev_miss, prev_down=prev_down, prev_opt=prev_opt)
        build_lldp_sheet(wb_out, "Downlinks",  down_rows, TAB_DOWN, prev_miss=prev_miss, prev_down=prev_down, prev_opt=prev_opt, is_downlinks=True)
        print(f"  LLDP — Mismatches:{len(miss_rows)} Downlinks:{len(down_rows)}")
        # Store for summary tab
        _miss_rows = miss_rows; _down_rows = down_rows; _lldp_rows = lldp_rows

    # Pass prev_opt to optics loop
    prev_opt_for_loop = prev_opt

    # Build downlinks set for cross-reference (hostname+interface)
    downlink_set = set()
    if ws_lldp:
        act_if_col_lldp = find_col(ws_lldp, 'Act. Interface', 'Act.Interface')
        host_col_lldp   = find_col(ws_lldp, 'Hostname')
        iface_col_lldp  = find_col(ws_lldp, 'Interface')
        if act_if_col_lldp and host_col_lldp and iface_col_lldp:
            for row in range(2, ws_lldp.max_row + 1):
                act_if = str(ws_lldp.cell(row, act_if_col_lldp).value or '').strip().lower()
                if act_if == 'interface down':
                    h = str(ws_lldp.cell(row, host_col_lldp).value or '').strip()
                    i = str(ws_lldp.cell(row, iface_col_lldp).value or '').strip()
                    if h and i:
                        downlink_set.add((h, i))

    # Detect optics format: JBP19 has Channel+Measured(dBm) cols
    import re as _ore
    _opt_hdr = {str(ws_optics.cell(1,c).value or '').strip(): c
                for c in range(1, ws_optics.max_column+1)} if ws_optics else {}
    optics_new_fmt = 'Measured (dBm)' in _opt_hdr

    if ws_optics and optics_new_fmt:
        # JBP19 optics: reformat to match JBP15 style
        # Interface | L&R | Rack | Elevation | Channel | Measured (dBm) |
        # Source_port | DMARC1 | Destination_port | Z Interface | Z L&R | Z Rack | Z Elevation | DL Flag | History
        ws_tab = wb_out.create_sheet("Optics")
        ws_tab.sheet_properties.tabColor = TAB_OPT
        _hdrs = [
            ("Interface",      HDR_BG,   14),
            ("L&R",            HDR_BG,    6),
            ("Rack",           HDR_BG,    8),
            ("Elevation",      HDR_BG,    6),
            ("Channel",        HDR_BG,    8),
            ("Measured (dBm)", HDR_BG,   14),
            ("T0 PP",          "C0504D", 30),
            ("T1 PP",          "17375E", 30),
            ("Z Interface",    "17375E", 12),
            ("Z L&R",          "17375E",  6),
            ("Z Rack",         "17375E",  8),
            ("Z Elevation",    "17375E",  6),
            ("DL Flag",        "595959", 22),
            ("History",        "595959", 22),
        ]
        for _ci, (_lbl, _bg, _w) in enumerate(_hdrs, 1):
            _c = ws_tab.cell(1, _ci); _c.value = _lbl
            _c.fill = fill(_bg); _c.font = font(HDR_FG, bold=True, sz=9); _c.alignment = center()
            ws_tab.column_dimensions[get_column_letter(_ci)].width = _w
        ws_tab.row_dimensions[1].height = 20
        _c_host  = _opt_hdr.get('Hostname', 2);    _c_rack  = _opt_hdr.get('Rack', 4)
        _c_elev  = _opt_hdr.get('Elevation', 5);   _c_trans = _opt_hdr.get('Transceiver', 6)
        _c_chan  = _opt_hdr.get('Channel', 7)
        _c_meas  = _opt_hdr.get('Measured (dBm)', 9)

        # Build T0->T1 reverse map from t1_rev values (device_a = "t0host t0iface")
        _t0_to_t1 = {}
        for (_t1h, _t1i), _tv in t1_rev.items():
            _da = _tv.get('device_a', '')
            if ' ' in _da:
                _t0h, _t0i = _da.split()[0], _da.split()[1]
                _t0_to_t1[(_t0h, _t0i)] = {
                    't1_host': _t1h, 't1_iface': _t1i,
                    't1_rack': _tv.get('rack_a', '')
                }

        _out_row = 2
        for _r in range(2, ws_optics.max_row + 1):
            _host  = str(ws_optics.cell(_r, _c_host).value  or '').strip()
            if not _host: continue
            _trans = str(ws_optics.cell(_r, _c_trans).value or '').strip()
            _rack  = str(ws_optics.cell(_r, _c_rack).value  or '').strip()
            _elev  = str(ws_optics.cell(_r, _c_elev).value  or '').strip()
            _chan  = ws_optics.cell(_r, _c_chan).value
            _meas  = ws_optics.cell(_r, _c_meas).value

            # Derive swp interface from transceiver number, try all sub-lanes
            _tn = _ore.search(r'transceiver(\d+)', _trans)
            _swp = None; _lr = ''; _pp = {}
            if _tn:
                _sb = 'swp' + _tn.group(1)
                for _lane in range(4):
                    _cand = f"{_sb}s{_lane}"
                    _pp_c = _cutsheet_pp.get((_host, _cand), {})
                    if _pp_c.get('source_port'):
                        _swp = _cand; _lr = phys_t0.get((_host, _cand), ''); _pp = _pp_c; break
                if not _swp:
                    _swp = _sb + 's0'; _lr = phys_t0.get((_host, _swp), '')

            # Z side — use T0->T1 reverse map
            _z_if = ''; _z_lr = ''; _z_rack = ''; _z_elev = ''
            if _swp:
                _t1m = _t0_to_t1.get((_host, _swp), {})
                if _t1m:
                    _z_if  = _t1m['t1_iface']
                    _z_lr  = phys_t1.get((_host, _swp), '')
                    _zrf   = _t1m['t1_rack']
                    _zrm   = _ore.match(r'Rack (\d+) U(\d+)', _zrf)
                    _z_rack = _zrm.group(1) if _zrm else _zrf.replace('Rack ', '')
                    _z_elev = _zrm.group(2) if _zrm else ''

            _is_dl = bool(_swp and (_host, _swp) in downlink_set)
            if _is_dl: continue  # skip — already shown as downlink
            _row_bg = "FFFFFF"
            _hist_flag, _hist_col = get_history_flag(_host, _swp or _trans, 'optic', prev_miss, prev_down, prev_opt)

            # 4 PP cols to match JBP15 format (DMARC2 blank for JBP19 2-hop paths)
            _vals = [_swp or _trans, _lr, _rack, _elev, _chan, _meas,
                     _pp.get('source_port', ''), _pp.get('dmarc1', ''),
                     _z_if, _z_lr, _z_rack, _z_elev]
            _bgs  = [_row_bg, "FFFFFF", _row_bg, _row_bg, _row_bg, _row_bg,
                     "FFFFFF", "FFFFFF",
                     "D9EAF7", "D9EAF7", "D9EAF7", "D9EAF7"]
            for _ci, (_v, _bg) in enumerate(zip(_vals, _bgs), 1):
                _c = ws_tab.cell(_out_row, _ci); _c.value = _v; _c.fill = fill(_bg)
                _c.font = font("000000", sz=9); _c.alignment = center()
            # DL Flag col 13
            _cf = ws_tab.cell(_out_row, 13); _cf.fill = fill("FFFFFF"); _cf.alignment = center()
            # History col 15
            _ch = ws_tab.cell(_out_row, 14)
            if _hist_flag:
                _ch.value = _hist_flag; _ch.fill = fill(_hist_col)
                _ch.font = Font(bold=True, color=WHITE, name="Arial", size=9)
            else:
                _ch.fill = fill("FFFFFF")
            _ch.alignment = center()
            ws_tab.row_dimensions[_out_row].height = 15
            _out_row += 1

        # Draw pair borders grouped by L&R (col 2)
        from openpyxl.styles import Border, Side
        _thin  = Side(style="thin",   color="AAAAAA")
        _thick = Side(style="medium", color="555555")
        _dr = 2
        while _dr <= ws_tab.max_row:
            _lr_val = ws_tab.cell(_dr, 2).value
            _ge = _dr
            while _ge + 1 <= ws_tab.max_row and ws_tab.cell(_ge+1, 2).value == _lr_val and _lr_val:
                _ge += 1
            for _rr in range(_dr, _ge + 1):
                _it = (_rr == _dr); _ib = (_rr == _ge)
                for _cc in range(1, ws_tab.max_column + 1):
                    ws_tab.cell(_rr, _cc).border = Border(
                        top    = _thick if _it else _thin,
                        bottom = _thick if _ib else Side(style=None),
                        left   = _thick if _cc == 1 else _thin,
                        right  = _thick if _cc == ws_tab.max_column else _thin,
                    )
            _dr = _ge + 1

    # JBP19 FEC — same clean format as Optics tab
    _fec_hdr = {str(ws_fec.cell(1,c).value or '').strip(): c
                for c in range(1, ws_fec.max_column+1)} if ws_fec else {}
    fec_new_fmt = any('BER' in k.upper() for k in _fec_hdr)

    if ws_fec and fec_new_fmt:
        ws_ftab = wb_out.create_sheet("FEC Errors")
        ws_ftab.sheet_properties.tabColor = TAB_FEC
        _fhdrs = [
            ("Interface",    HDR_BG,    14),
            ("L&R",          HDR_BG,     6),
            ("Rack",         HDR_BG,     8),
            ("Elevation",    HDR_BG,     6),
            ("Lock Status",  HDR_BG,    28),
            ("Pre-FEC BER",  HDR_BG,    12),
            ("T0 PP",        "C0504D",  30),
            ("T1 PP",        "17375E",  30),
            ("Z Interface",  "17375E",  12),
            ("Z L&R",        "17375E",   6),
            ("Z Rack",       "17375E",   8),
            ("Z Elevation",  "17375E",   6),
            ("Severity",     "595959",  12),
            ("History",      "595959",  22),
        ]
        for _ci, (_lbl, _bg, _w) in enumerate(_fhdrs, 1):
            _c = ws_ftab.cell(1, _ci); _c.value = _lbl
            _c.fill = fill(_bg); _c.font = font(HDR_FG, bold=True, sz=9); _c.alignment = center()
            ws_ftab.column_dimensions[get_column_letter(_ci)].width = _w
        ws_ftab.row_dimensions[1].height = 20

        # Detect column positions
        _fc_host  = _fec_hdr.get('Hostname', 2)
        _fc_iface = _fec_hdr.get('Interface', 3)
        _fc_rack  = _fec_hdr.get('Rack', 4)
        _fc_elev  = _fec_hdr.get('Elevation', 5)
        _fc_lock  = _fec_hdr.get('Lock Status', 6)
        _fc_ber   = next((v for k,v in _fec_hdr.items() if 'BER' in k.upper()), 7)

        # Build T0->T1 map if not already built
        if '_t0_to_t1' not in dir():
            _t0_to_t1 = {}
            for (_t1h, _t1i), _tv in t1_rev.items():
                _da = _tv.get('device_a', '')
                if ' ' in _da:
                    _t0h, _t0i = _da.split()[0], _da.split()[1]
                    _t0_to_t1[(_t0h, _t0i)] = {
                        't1_host': _t1h, 't1_iface': _t1i,
                        't1_rack': _tv.get('rack_a', '')
                    }

        _fout_row = 2
        for _fr in range(2, ws_fec.max_row + 1):
            _fhost  = str(ws_fec.cell(_fr, _fc_host).value  or '').strip()
            if not _fhost: continue
            _fiface = str(ws_fec.cell(_fr, _fc_iface).value or '').strip()
            _frack  = str(ws_fec.cell(_fr, _fc_rack).value  or '').strip()
            _felev  = str(ws_fec.cell(_fr, _fc_elev).value  or '').strip()
            _flock  = str(ws_fec.cell(_fr, _fc_lock).value  or '').strip()
            _fber   = ws_fec.cell(_fr, _fc_ber).value

            _flr  = phys_t0.get((_fhost, _fiface), '')
            _fpp  = _cutsheet_pp.get((_fhost, _fiface), {})
            _is_p = _fiface.endswith('s0') or _fiface.endswith('s2')
            _frow_bg = "FFFFFF" if _is_p else LOG_BG

            _fz_if = ''; _fz_lr = ''; _fz_rack = ''; _fz_elev = ''
            _ft1m = _t0_to_t1.get((_fhost, _fiface), {})
            if _ft1m:
                _fz_if  = _ft1m['t1_iface']
                _fz_lr  = phys_t1.get((_fhost, _fiface), '')
                _fzrf   = _ft1m['t1_rack']
                _fzrm   = _ore.match(r'Rack (\d+) U(\d+)', _fzrf)
                _fz_rack = _fzrm.group(1) if _fzrm else _fzrf.replace('Rack ', '')
                _fz_elev = _fzrm.group(2) if _fzrm else ''

            _fhist_flag, _fhist_col = get_history_flag(_fhost, _fiface, 'fec', prev_miss, prev_down, prev_opt)

            _fvals = [_fiface, _flr, _frack, _felev, _flock, _fber,
                      _fpp.get('source_port', ''), _fpp.get('dmarc1', ''),
                      _fz_if, _fz_lr, _fz_rack, _fz_elev]
            _fbgs  = [_frow_bg, "FFFFFF", _frow_bg, _frow_bg, _frow_bg, _frow_bg,
                      "FFFFFF", "FFFFFF", "D9EAF7", "D9EAF7", "D9EAF7", "D9EAF7"]
            for _ci, (_v, _bg) in enumerate(zip(_fvals, _fbgs), 1):
                _c = ws_ftab.cell(_fout_row, _ci); _c.value = _v; _c.fill = fill(_bg)
                _c.font = font("000000", sz=9); _c.alignment = center()
            _fcf = ws_ftab.cell(_fout_row, 13); _fcf.value = ''; _fcf.fill = fill("FFFFFF"); _fcf.alignment = center()
            _fch = ws_ftab.cell(_fout_row, 14)
            if _fhist_flag:
                _fch.value = _fhist_flag; _fch.fill = fill(_fhist_col)
                _fch.font = Font(bold=True, color=WHITE, name="Arial", size=9)
            else:
                _fch.fill = fill("FFFFFF")
            _fch.alignment = center()
            ws_ftab.row_dimensions[_fout_row].height = 15
            _fout_row += 1

        # Borders grouped by L&R (col 2)
        from openpyxl.styles import Border, Side
        _fthin = Side(style="thin", color="AAAAAA"); _fthick = Side(style="medium", color="555555")
        _fdr = 2
        while _fdr <= ws_ftab.max_row:
            _flv = ws_ftab.cell(_fdr, 2).value; _fge = _fdr
            while _fge + 1 <= ws_ftab.max_row and ws_ftab.cell(_fge+1, 2).value == _flv and _flv:
                _fge += 1
            for _frr in range(_fdr, _fge + 1):
                _fit = (_frr == _fdr); _fib = (_frr == _fge)
                for _fcc in range(1, ws_ftab.max_column + 1):
                    ws_ftab.cell(_frr, _fcc).border = Border(
                        top    = _fthick if _fit else _fthin,
                        bottom = _fthick if _fib else Side(style=None),
                        left   = _fthick if _fcc == 1 else _fthin,
                        right  = _fthick if _fcc == ws_ftab.max_column else _fthin,
                    )
            _fdr = _fge + 1

        # Optics and FEC — physical highlighting + T0 L&R + T1 L&R + blank PP_info
    for ws_extra, tab_name, tab_col in [
        (ws_optics if not optics_new_fmt else None, "Optics", TAB_OPT),
        (ws_fec if not fec_new_fmt else None, "FEC Errors", TAB_FEC),
    ]:
        current_type_for_tab = 'optic' 
        if not ws_extra: continue
        host_col    = find_col(ws_extra, 'Hostname')
        iface_col   = find_col(ws_extra, 'Interface') or find_col(ws_extra, 'Transceiver')
        z_iface_col = find_col(ws_extra, 'Z Interface')
        if not host_col or not iface_col: continue

        ws_out = wb_out.create_sheet(tab_name)
        ws_out.sheet_properties.tabColor = tab_col
        ncols = ws_extra.max_column

        # Two inserted cols: T0 L&R after Interface, T1 L&R after Z Interface
        t0_lr_col = iface_col + 1         # inserted col 1
        # Z Interface in source = z_iface_col, in output = z_iface_col + 1 (shifted by t0_lr insertion)
        t1_lr_col = (z_iface_col + 1) + 1 if z_iface_col else None  # inserted col 2
        total_out_cols = ncols + (2 if t1_lr_col else 1)

        def src_to_out(src_col):
            """Map source col index to output col index accounting for insertions"""
            out = src_col
            if src_col >= t0_lr_col: out += 1
            if t1_lr_col and src_col >= (z_iface_col): out += 1
            return out

        # Build header row
        insert_cols = {t0_lr_col: ("L&R", HDR_BG)}
        if t1_lr_col:
            insert_cols[t1_lr_col] = ("Z L&R", HDR_BG)

        flag_col = total_out_cols + 1

        out_col = 1
        src_col = 1
        while out_col <= total_out_cols:
            c = ws_out.cell(1, out_col)
            if out_col in insert_cols:
                label, bg = insert_cols[out_col]
                c.value = label
                ws_out.column_dimensions[get_column_letter(out_col)].width = 6
            elif src_col <= ncols:
                c.value = ws_extra.cell(1, src_col).value
                src_col += 1
            c.fill = fill(HDR_BG); c.font = font(HDR_FG, bold=True, sz=9); c.alignment = center()
            out_col += 1
        cf = ws_out.cell(1, flag_col)
        cf.value = "DL Flag"; cf.fill = fill("595959")
        cf.font = Font(bold=True, color=WHITE, name="Arial", size=8)
        cf.alignment = Alignment(horizontal="center", vertical="center")
        ws_out.column_dimensions[get_column_letter(flag_col)].width = 24
        # History column
        hist_col_num = flag_col + 1
        ch = ws_out.cell(1, hist_col_num)
        ch.value = "History"; ch.fill = fill("595959")
        ch.font = Font(bold=True, color=WHITE, name="Arial", size=8)
        ch.alignment = Alignment(horizontal="center", vertical="center")
        ws_out.column_dimensions[get_column_letter(hist_col_num)].width = 22
        ws_out.row_dimensions[1].height = 20

        dl_overlap = 0
        # First pass — build (host, iface) -> row number lookup for partner alignment
        raw_opt_rows = {}
        for _r in range(2, ws_extra.max_row + 1):
            _h = str(ws_extra.cell(_r, host_col).value or '').strip()
            _i = str(ws_extra.cell(_r, iface_col).value or '').strip()
            if _h and _i: raw_opt_rows[(_h, _i)] = _r

        # Find patch panel cols in this sheet (Source_port, DMARC1, DMARC2, Destination_port)
        pp_col_names = ['Source_port', 'DMARC1', 'DMARC2', 'Destination_port']
        pp_cols = []
        for name in pp_col_names:
            c = find_col(ws_extra, name)
            if c: pp_cols.append(c)

        out_row = 2
        for row in range(2, ws_extra.max_row + 1):
            host  = str(ws_extra.cell(row, host_col).value or '').strip()
            iface = str(ws_extra.cell(row, iface_col).value or '').strip()
            if not host or not iface: continue
            t0_lbl, t1_lbl, is_p = get_labels(host, iface, phys_t0, phys_t1)

            # For logical rows — find physical partner and copy its patch panel cols
            partner_pp_override = {}
            if not is_p:
                m_pp = re.match(r'(swp\d+)s(\d+)', iface)
                if m_pp:
                    base_pp, lane_pp = m_pp.group(1), int(m_pp.group(2))
                    partner_lane_pp = {0:1,1:0,2:3,3:2}.get(lane_pp)
                    if partner_lane_pp is not None:
                        partner_iface_pp = f"{base_pp}s{partner_lane_pp}"
                        partner_row_pp = raw_opt_rows.get((host, partner_iface_pp))
                        if partner_row_pp:
                            # Partner in report — copy directly
                            for pp_c in pp_cols:
                                partner_pp_override[pp_c] = ws_extra.cell(partner_row_pp, pp_c).value
                        else:
                            # Partner not in report — fall back to cutsheet PP data
                            pp_fb = _cutsheet_pp.get((host, partner_iface_pp), {})
                            if pp_fb:
                                pp_key_map = {
                                    'source_port': 'Source_port',
                                    'dmarc1':      'DMARC1',
                                    'dmarc2':      'DMARC2',
                                    'dest_port':   'Destination_port',
                                }
                                for cs_key, col_name in pp_key_map.items():
                                    c_num = find_col(ws_extra, col_name)
                                    if c_num:
                                        partner_pp_override[c_num] = pp_fb.get(cs_key, "")
            is_also_downlink = (host, iface) in downlink_set
            if is_also_downlink: dl_overlap += 1

            row_bg = "C8C8C8" if is_also_downlink else ("FFFFFF" if is_p else LOG_BG)
            lr_bg  = "A8A8A8" if is_also_downlink else (LR_BG if is_p else LR_LOG)
            txt_fg = "888888" if is_also_downlink else "000000"

            out_col = 1; src_col = 1
            while out_col <= total_out_cols:
                c = ws_out.cell(out_row, out_col)
                if out_col == t0_lr_col:
                    c.value = t0_lbl
                    c.fill = fill(lr_bg)
                    c.font = font(sz=8, bold=True, color=txt_fg); c.alignment = center()
                elif t1_lr_col and out_col == t1_lr_col:
                    c.value = t1_lbl
                    c.fill = fill(lr_bg)
                    c.font = font(sz=8, bold=True, color=txt_fg); c.alignment = center()
                elif src_col <= ncols:
                    # Use partner's patch panel data for logical rows
                    raw = partner_pp_override.get(src_col, ws_extra.cell(row, src_col).value)
                    c.value = '' if str(raw or '').startswith('PP_info') else raw
                    c.fill = fill(row_bg)
                    c.font = font(sz=8, color=txt_fg); c.alignment = vcenter()
                    src_col += 1
                out_col += 1

            cf = ws_out.cell(out_row, flag_col)
            if is_also_downlink:
                cf.value = "⬇️ Also Downlink — skip"
                cf.fill  = fill("C8C8C8")
                cf.font  = Font(bold=True, color="666666", name="Arial", size=8)
            else:
                cf.fill = fill(row_bg); cf.font = font(sz=8)
            cf.alignment = Alignment(horizontal="center", vertical="center")

            # History flag
            hist_text, hist_colour = get_history_flag(
                host, iface, current_type_for_tab, prev_miss, prev_down, prev_opt_for_loop
            )
            ch = ws_out.cell(out_row, hist_col_num)
            if hist_text:
                ch.value = hist_text
                ch.fill  = fill(hist_colour)
                ch.font  = Font(bold=True, color="FFFFFF", name="Arial", size=8)
            else:
                ch.fill = fill(row_bg); ch.font = font(sz=8)
            ch.alignment = Alignment(horizontal="center", vertical="center")

            ws_out.row_dimensions[out_row].height = 15
            out_row += 1

        ws_out.freeze_panes = "A2"

        # ── Borders around each physical+logical pair ─────────────────────────
        from openpyxl.styles import Border, Side
        thin2  = Side(style="thin",   color="AAAAAA")
        thick2 = Side(style="medium", color="555555")
        lr_col2 = t0_lr_col  # L&R col position
        dr = 2
        while dr <= ws_out.max_row:
            lr_val2 = ws_out.cell(dr, lr_col2).value
            grp_end = dr
            while (grp_end + 1 <= ws_out.max_row and
                   ws_out.cell(grp_end+1, lr_col2).value == lr_val2 and lr_val2):
                grp_end += 1
            for rr in range(dr, grp_end + 1):
                is_top2    = (rr == dr)
                is_bottom2 = (rr == grp_end)
                for cc in range(1, ws_out.max_column + 1):
                    ws_out.cell(rr, cc).border = Border(
                        top    = thick2 if is_top2    else thin2,
                        bottom = thick2 if is_bottom2 else Side(style=None),
                        left   = thick2 if cc == 1    else thin2,
                        right  = thick2 if cc == ws_out.max_column else thin2,
                    )
            dr = grp_end + 1

        print(f"  {tab_name} — {out_row-2} rows | {dl_overlap} flagged as also-downlink")

    # ── Build Summary tab ────────────────────────────────────────────────────
    if ws_lldp:
        # Build current optics by rack for summary
        curr_opt_rack = {}
        if ws_optics:
            _ohc = find_col(ws_optics, 'Hostname')
            _oic = find_col(ws_optics, 'Interface') or find_col(ws_optics, 'Transceiver')
            _orc = find_col(ws_optics, 'Rack')
            if _ohc and _oic:
                for _r in range(2, ws_optics.max_row+1):
                    _h = str(ws_optics.cell(_r, _ohc).value or '').strip()
                    _i = str(ws_optics.cell(_r, _oic).value or '').strip()
                    _rack = str(ws_optics.cell(_r, _orc).value or '').strip() if _orc else 'Unknown'
                    if _h and _i:
                        if _rack not in curr_opt_rack: curr_opt_rack[_rack] = set()
                        curr_opt_rack[_rack].add((_h, _i))

        build_summary_tab(wb_out, _lldp_rows, _miss_rows, _down_rows,
                          prev_miss, prev_down, prev_opt,
                          os.path.basename(report_path),
                          os.path.basename(prev_report_path) if prev_report_path else None,
                          prev_rack_map=prev_rack_map, prev_opt_rack_map=prev_opt_rack_map,
                          curr_opt_rack=curr_opt_rack)

    base, ext = os.path.splitext(report_path)
    out_path  = base + "_highlighted" + ext
    wb_out.save(out_path)

    msg = (f"Done!\n\n"
           f"All: {len(lldp_rows)}  |  Mispatches: {len(miss_rows)}  |  Downlinks: {len(down_rows)}\n\n"
           f"Saved to:\n{out_path}")
    print(f"\n{msg}")
    show_msg("Complete ✅", msg)

    try:
        import subprocess
        if sys.platform   == "win32":  os.startfile(out_path)
        elif sys.platform == "darwin": subprocess.run(["open", out_path])
        else:                          subprocess.run(["xdg-open", out_path])
    except: pass

if __name__ == "__main__":
    main()
